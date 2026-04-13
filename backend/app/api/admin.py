import hashlib
import io
import os
import xml.etree.ElementTree as ET
from datetime import date, datetime
from app.core.timezone import now_kst
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
from sqlalchemy import text, func
from typing import Optional, List
from app.core.database import get_db
from app.models.user import User
from app.models.problem import Problem
from app.models.solution import Solution, RESULT_CODES, LANGUAGE_NAMES
from app.models.classroom import Classroom, ClassMember
from app.api.deps import get_current_admin, get_current_teacher

router = APIRouter(prefix="/admin", tags=["admin"])


def md5(text: str) -> str:
    return hashlib.md5(text.encode()).hexdigest()


# ─── Dashboard Stats ────────────────────────────────────────────────────────

@router.get("/stats")
def get_stats(db: Session = Depends(get_db), _=Depends(get_current_admin)):
    total_users = db.query(func.count(User.user_id)).filter(User.defunct == "N").scalar()
    total_problems = db.query(func.count(Problem.problem_id)).filter(Problem.defunct == "N").scalar()
    total_submissions = db.query(func.count(Solution.solution_id)).scalar()
    today_submissions = db.query(func.count(Solution.solution_id)).filter(
        func.date(Solution.in_date) == date.today()
    ).scalar()
    accepted = db.query(func.count(Solution.solution_id)).filter(Solution.result == 4).scalar()

    recent = (
        db.query(Solution)
        .order_by(Solution.solution_id.desc())
        .limit(10)
        .all()
    )
    recent_list = [
        {
            "solution_id": s.solution_id,
            "user_id": s.user_id,
            "problem_id": s.problem_id,
            "result": RESULT_CODES.get(s.result, str(s.result)),
            "result_code": s.result,
            "language": LANGUAGE_NAMES.get(s.language, str(s.language)),
            "in_date": s.in_date.isoformat() if s.in_date else None,
        }
        for s in recent
    ]

    return {
        "total_users": total_users,
        "total_problems": total_problems,
        "total_submissions": total_submissions,
        "today_submissions": today_submissions,
        "accepted": accepted,
        "recent_submissions": recent_list,
    }


# ─── Problem Management ──────────────────────────────────────────────────────

@router.get("/problems")
def list_problems(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    query = db.query(Problem)
    if search:
        query = query.filter(
            Problem.title.ilike(f"%{search}%") |
            Problem.problem_id.cast(str).ilike(f"%{search}%")
        )
    total = query.count()
    problems = query.order_by(Problem.problem_id.desc()).offset((page - 1) * page_size).limit(page_size).all()
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "problems": [
            {
                "problem_id": p.problem_id,
                "title": p.title,
                "submit": p.submit,
                "accepted": p.accepted,
                "defunct": p.defunct,
                "in_date": p.in_date.isoformat() if p.in_date else None,
            }
            for p in problems
        ],
    }


# ─── User Management ─────────────────────────────────────────────────────────

@router.get("/users")
def list_users(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    show_defunct: bool = Query(False),
    role: Optional[str] = Query(None),        # "student" | "staff"
    grade_filter: Optional[int] = Query(None, alias="grade"),
    class_num_filter: Optional[int] = Query(None, alias="class_num"),
    sort_by: str = Query("reg_time"),          # reg_time | user_id | nick | submit | grade | class_num | student_num
    sort_order: str = Query("desc"),           # asc | desc
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    from sqlalchemy import asc, desc as sqldesc, case as sqcase

    # ── privilege 로딩 ──────────────────────────────────────────────────
    priv_rows = db.execute(
        text("SELECT user_id, rightstr FROM privilege WHERE rightstr IN ('administrator','teacher') AND defunct='N'")
    ).fetchall()
    admin_ids  = {r[0] for r in priv_rows if r[1] == "administrator"}
    teacher_ids = {r[0] for r in priv_rows if r[1] == "teacher"}
    staff_ids  = admin_ids | teacher_ids

    # ── class_members 서브쿼리 (유저당 최신 1건) ───────────────────────
    cm_latest = db.query(
        ClassMember.user_id,
        func.max(ClassMember.id).label("max_id"),
    ).group_by(ClassMember.user_id).subquery()

    cm_sub = db.query(ClassMember).join(
        cm_latest,
        (ClassMember.user_id == cm_latest.c.user_id) &
        (ClassMember.id == cm_latest.c.max_id),
    ).subquery()

    # ── 기본 쿼리 ──────────────────────────────────────────────────────
    query = db.query(
        User,
        cm_sub.c.grade,
        cm_sub.c.class_num,
        cm_sub.c.student_num,
    ).outerjoin(cm_sub, User.user_id == cm_sub.c.user_id)

    if not show_defunct:
        query = query.filter(User.defunct == "N")

    # role 필터
    if role == "student":
        if staff_ids:
            query = query.filter(~User.user_id.in_(staff_ids))
    elif role == "staff":
        if staff_ids:
            query = query.filter(User.user_id.in_(staff_ids))
        else:
            query = query.filter(text("1=0"))

    # 검색 (아이디 / 이름 / 학교)
    if search:
        query = query.filter(
            User.user_id.ilike(f"%{search}%") |
            User.nick.ilike(f"%{search}%") |
            User.school.ilike(f"%{search}%")
        )

    # 학년 / 반 필터
    if grade_filter is not None:
        query = query.filter(cm_sub.c.grade == grade_filter)
    if class_num_filter is not None:
        query = query.filter(cm_sub.c.class_num == class_num_filter)

    total = query.count()

    # 정렬
    sort_map = {
        "user_id":     User.user_id,
        "nick":        User.nick,
        "submit":      User.submit,
        "grade":       cm_sub.c.grade,
        "class_num":   cm_sub.c.class_num,
        "student_num": cm_sub.c.student_num,
        "reg_time":    User.reg_time,
    }
    col = sort_map.get(sort_by, User.reg_time)
    # NULL을 마지막으로: NULL이면 1, 아니면 0 으로 먼저 정렬
    null_last = sqcase((col == None, 1), else_=0)  # noqa: E711
    dir_fn = asc if sort_order == "asc" else sqldesc
    query = query.order_by(null_last, dir_fn(col))

    rows = query.offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "users": [
            {
                "user_id": u.user_id,
                "nick": u.nick,
                "school": u.school,
                "submit": u.submit,
                "solved": u.solved,
                "defunct": u.defunct,
                "reg_time": u.reg_time.isoformat() if u.reg_time else None,
                "grade": g,
                "class_num": cn,
                "student_num": sn,
                "is_admin": u.user_id in admin_ids,
                "is_teacher": u.user_id in teacher_ids or u.user_id in admin_ids,
            }
            for u, g, cn, sn in rows
        ],
    }


@router.get("/users/import-template")
def download_template(_=Depends(get_current_teacher)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "사용자 일괄 등록"

    # 컬럼: 학교, 학년, 반, 번호, 아이디*, 비밀번호*, 이름, 별명, 비고
    headers =    ["학교",  "학년", "반", "번호", "아이디*", "비밀번호*", "이름",  "별명",  "비고"]
    col_widths = [18,      8,     6,   6,      14,        14,          14,     14,     20]

    required_fill = PatternFill("solid", fgColor="C0392B")
    normal_fill   = PatternFill("solid", fgColor="2E86C1")

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = required_fill if h.endswith("*") else normal_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    examples = [
        ["○○중학교", 2, 3, 1, "student01", "pass1234", "홍길동", "길동이", ""],
        ["○○중학교", 2, 3, 2, "student02", "pass1234", "이순신", "",       ""],
        ["○○고등학교", 1, 1, 5, "student03", "pass1234", "강감찬", "감찬",  "전학생"],
    ]
    for row_data in examples:
        ws.append(row_data)

    ws2 = wb.create_sheet("작성 안내")
    for row in [
        ["컬럼",     "설명"],
        ["학교",     "학교/소속 (선택)"],
        ["학년",     "숫자 (선택)"],
        ["반",       "숫자 (선택)"],
        ["번호",     "숫자 (선택)"],
        ["아이디*",  "로그인 ID, 필수, 영문/숫자"],
        ["비밀번호*","4자 이상, 필수"],
        ["이름",     "실명 (미입력 시 아이디로 대체)"],
        ["별명",     "표시 닉네임 (미입력 시 이름으로 대체)"],
        ["비고",     "참고 메모 (등록에 사용되지 않음)"],
    ]:
        ws2.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"},
    )


@router.post("/users/bulk-import")
async def bulk_import_users(
    file: UploadFile = File(...),
    class_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher),
):
    import openpyxl

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail=".xlsx 파일만 지원합니다")

    # 학급 존재 확인
    cls = None
    if class_id:
        cls = db.query(Classroom).filter(Classroom.id == class_id, Classroom.archived == 0).first()
        if not cls:
            raise HTTPException(status_code=404, detail="학급을 찾을 수 없습니다")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="파일을 읽을 수 없습니다")

    results = []
    skipped = []

    def _to_int(val) -> Optional[int]:
        """엑셀 숫자(int/float)·문자열 모두 안전하게 정수 변환."""
        if val is None:
            return None
        if isinstance(val, (int, float)):
            return int(val)
        try:
            return int(float(str(val).strip()))
        except (ValueError, TypeError):
            return None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[4]:   # 아이디 없으면 skip
            continue

        school      = str(row[0]).strip() if row[0] else ""
        grade       = _to_int(row[1])
        class_num   = _to_int(row[2])
        student_num = _to_int(row[3])
        user_id     = str(row[4]).strip()
        password    = str(row[5]).strip() if row[5] else ""
        real_name   = str(row[6]).strip() if row[6] else ""
        nickname    = str(row[7]).strip() if row[7] else ""
        # row[8] = 비고, 무시

        if not password:
            skipped.append({"row": row_idx, "user_id": user_id, "reason": "비밀번호 누락"})
            continue
        if len(password) < 4:
            skipped.append({"row": row_idx, "user_id": user_id, "reason": "비밀번호 4자 미만"})
            continue
        if db.query(User).filter(User.user_id == user_id).first():
            skipped.append({"row": row_idx, "user_id": user_id, "reason": "이미 존재하는 아이디"})
            continue

        # 별명 > 이름 > 아이디 순으로 닉네임 결정
        display = nickname or real_name or user_id

        db.add(User(
            user_id=user_id,
            password=md5(password),
            nick=display,
            school=school,
            email=None,
            ip="",
            reg_time=now_kst(),
            submit=0,
            solved=0,
        ))
        db.flush()

        # 학년/반/번호 또는 학급 배정이 있으면 ClassMember 생성
        # class_id 없이도 grade/class_num/student_num 저장 가능 (NULL 허용)
        if cls or grade or class_num or student_num:
            db.add(ClassMember(
                class_id=cls.id if cls else None,
                user_id=user_id,
                role="student",
                grade=grade,
                class_num=class_num,
                student_num=student_num,
                joined_at=now_kst(),
            ))

        results.append({"row": row_idx, "user_id": user_id, "name": real_name or user_id})

    db.commit()
    return {"created": len(results), "skipped": len(skipped), "details": results, "errors": skipped}


@router.delete("/users")
def bulk_delete_users(
    user_ids: List[str] = Body(..., embed=True),
    permanent: bool = Query(False),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher),
):
    """선택된 사용자 일괄 삭제 (soft 또는 permanent)."""
    if not user_ids:
        raise HTTPException(status_code=400, detail="삭제할 사용자를 선택하세요")

    my_privs = {
        r[0] for r in db.execute(
            text("SELECT rightstr FROM privilege WHERE user_id=:uid AND defunct='N'"),
            {"uid": current_user.user_id},
        ).fetchall()
    }
    is_admin = "administrator" in my_privs

    if permanent and not is_admin:
        raise HTTPException(status_code=403, detail="영구 삭제는 관리자만 가능합니다")

    deleted, skipped = [], []
    for uid in user_ids:
        if uid == current_user.user_id:
            skipped.append({"user_id": uid, "reason": "자기 자신"})
            continue

        target = db.query(User).filter(User.user_id == uid).first()
        if not target:
            skipped.append({"user_id": uid, "reason": "존재하지 않음"})
            continue

        target_privs = {
            r[0] for r in db.execute(
                text("SELECT rightstr FROM privilege WHERE user_id=:uid AND defunct='N'"),
                {"uid": uid},
            ).fetchall()
        }

        if not is_admin and ("administrator" in target_privs or "teacher" in target_privs):
            skipped.append({"user_id": uid, "reason": "권한 없음 (교사/관리자)"})
            continue

        if permanent:
            if "administrator" in target_privs:
                skipped.append({"user_id": uid, "reason": "관리자 계정 영구삭제 불가"})
                continue
            db.execute(text("DELETE FROM privilege WHERE user_id=:uid"), {"uid": uid})
            db.execute(text("DELETE FROM class_members WHERE user_id=:uid"), {"uid": uid})
            db.delete(target)
        else:
            target.defunct = "Y"

        deleted.append(uid)

    db.commit()
    return {"deleted": len(deleted), "skipped": len(skipped), "skipped_details": skipped}


@router.delete("/users/{user_id}/permanent")
def delete_user_permanent(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_admin),
):
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="자기 자신은 삭제할 수 없습니다")

    target = db.query(User).filter(User.user_id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    target_privs = {
        r[0] for r in db.execute(
            text("SELECT rightstr FROM privilege WHERE user_id=:uid AND defunct='N'"),
            {"uid": user_id},
        ).fetchall()
    }
    if "administrator" in target_privs:
        raise HTTPException(status_code=403, detail="관리자 계정은 영구 삭제할 수 없습니다")

    # 관련 데이터 정리 후 계정 삭제
    db.execute(text("DELETE FROM privilege WHERE user_id=:uid"), {"uid": user_id})
    db.execute(text("DELETE FROM class_members WHERE user_id=:uid"), {"uid": user_id})
    db.delete(target)
    db.commit()
    return {"ok": True}


@router.patch("/users/{user_id}/defunct")
def toggle_user_defunct(
    user_id: str,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    if user_id == current_admin.user_id:
        raise HTTPException(status_code=400, detail="자기 자신은 비활성화할 수 없습니다")
    user = db.query(User).filter(User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    user.defunct = "N" if user.defunct == "Y" else "Y"
    db.commit()
    return {"user_id": user_id, "defunct": user.defunct}


@router.patch("/users/{user_id}/admin")
def toggle_user_admin(
    user_id: str,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    user = db.query(User).filter(User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    existing = db.execute(
        text("SELECT defunct FROM privilege WHERE user_id=:uid AND rightstr='administrator' LIMIT 1"),
        {"uid": user_id},
    ).fetchone()

    if existing:
        new_defunct = "N" if existing[0] == "Y" else "Y"
        db.execute(
            text("UPDATE privilege SET defunct=:d WHERE user_id=:uid AND rightstr='administrator'"),
            {"d": new_defunct, "uid": user_id},
        )
        is_admin = new_defunct == "N"
    else:
        db.execute(
            text("INSERT INTO privilege (user_id, rightstr, valuestr, defunct) VALUES (:uid, 'administrator', 'true', 'N')"),
            {"uid": user_id},
        )
        is_admin = True

    db.commit()
    return {"user_id": user_id, "is_admin": is_admin}


@router.patch("/users/{user_id}/teacher")
def toggle_user_teacher(
    user_id: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    user = db.query(User).filter(User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    existing = db.execute(
        text("SELECT defunct FROM privilege WHERE user_id=:uid AND rightstr='teacher' LIMIT 1"),
        {"uid": user_id},
    ).fetchone()

    if existing:
        new_defunct = "N" if existing[0] == "Y" else "Y"
        db.execute(
            text("UPDATE privilege SET defunct=:d WHERE user_id=:uid AND rightstr='teacher'"),
            {"d": new_defunct, "uid": user_id},
        )
        is_teacher = new_defunct == "N"
    else:
        db.execute(
            text("INSERT INTO privilege (user_id, rightstr, valuestr, defunct) VALUES (:uid, 'teacher', 'true', 'N')"),
            {"uid": user_id},
        )
        is_teacher = True

    db.commit()
    return {"user_id": user_id, "is_teacher": is_teacher}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher),
):
    if user_id == current_user.user_id:
        raise HTTPException(status_code=400, detail="자기 자신은 삭제할 수 없습니다")

    target = db.query(User).filter(User.user_id == user_id).first()
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    # 대상 권한 확인
    target_privs = {
        r[0] for r in db.execute(
            text("SELECT rightstr FROM privilege WHERE user_id=:uid AND defunct='N'"),
            {"uid": user_id},
        ).fetchall()
    }

    # 현재 사용자 권한 확인
    my_privs = {
        r[0] for r in db.execute(
            text("SELECT rightstr FROM privilege WHERE user_id=:uid AND defunct='N'"),
            {"uid": current_user.user_id},
        ).fetchall()
    }
    is_admin = "administrator" in my_privs

    # 교사는 관리자/교사 삭제 불가
    if not is_admin and ("administrator" in target_privs or "teacher" in target_privs):
        raise HTTPException(status_code=403, detail="교사는 학생만 삭제할 수 있습니다")

    target.defunct = "Y"
    db.commit()
    return {"ok": True}


@router.patch("/users/{user_id}/password")
def reset_user_password(
    user_id: str,
    body: dict,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    user = db.query(User).filter(User.user_id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    new_pw = body.get("password", "")
    if len(new_pw) < 4:
        raise HTTPException(status_code=400, detail="비밀번호는 4자 이상이어야 합니다")
    user.password = md5(new_pw)
    db.commit()
    return {"ok": True}


# ─── Submission Management ───────────────────────────────────────────────────

@router.get("/submissions")
def list_submissions(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    user_id: Optional[str] = None,
    problem_id: Optional[int] = None,
    result: Optional[int] = None,
    language: Optional[int] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    query = db.query(Solution)
    if user_id:
        query = query.filter(Solution.user_id.ilike(f"%{user_id}%"))
    if problem_id is not None:
        query = query.filter(Solution.problem_id == problem_id)
    if result is not None:
        query = query.filter(Solution.result == result)
    if language is not None:
        query = query.filter(Solution.language == language)

    total = query.count()
    submissions = query.order_by(Solution.solution_id.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "submissions": [
            {
                "solution_id": s.solution_id,
                "user_id": s.user_id,
                "problem_id": s.problem_id,
                "result": RESULT_CODES.get(s.result, str(s.result)),
                "result_code": s.result,
                "language": LANGUAGE_NAMES.get(s.language, str(s.language)),
                "time": s.time,
                "memory": s.memory,
                "code_length": s.code_length,
                "in_date": s.in_date.isoformat() if s.in_date else None,
            }
            for s in submissions
        ],
    }


# ─── XML Import / Export ─────────────────────────────────────────────────────

JUDGE_DATA_PATH = os.environ.get("JUDGE_DATA_PATH", "/judge_data")


def _read_testcases(problem_id: int) -> list[dict]:
    """judge_data 디렉토리에서 테스트 케이스 읽기"""
    data_dir = f"{JUDGE_DATA_PATH}/{problem_id}"
    tests = []
    if not os.path.exists(data_dir):
        return tests
    i = 1
    while True:
        inp = os.path.join(data_dir, f"{i}.in")
        out = os.path.join(data_dir, f"{i}.out")
        if not os.path.exists(inp):
            break
        tests.append({
            "input": open(inp, encoding="utf-8", errors="replace").read(),
            "output": open(out, encoding="utf-8", errors="replace").read() if os.path.exists(out) else "",
        })
        i += 1
    return tests


def _write_testcases(problem_id: int, tests: list[dict]):
    """테스트 케이스를 judge_data 디렉토리에 저장"""
    data_dir = f"{JUDGE_DATA_PATH}/{problem_id}"
    os.makedirs(data_dir, exist_ok=True)
    for i, t in enumerate(tests, 1):
        with open(os.path.join(data_dir, f"{i}.in"), "w", encoding="utf-8", newline="\n") as f:
            f.write(t["input"])
        with open(os.path.join(data_dir, f"{i}.out"), "w", encoding="utf-8", newline="\n") as f:
            f.write(t["output"])


def _build_xml(problems_with_tests: list[dict]) -> str:
    """CSLfps XML 생성"""
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<!-- exported by HustOJ Evolution -->',
        '<CSLfps version="1.1" url="https://github.com/melongist/CSL/tree/master/HUSTOJ">',
        '  <generator name="HustOJ-Evolution" url=""/>',
    ]
    for pw in problems_with_tests:
        p = pw["problem"]
        tests = pw["tests"]
        tl = float(p.time_limit) if p.time_limit else 1.0
        ml = int(p.memory_limit) if p.memory_limit else 128

        def cdata(text: str) -> str:
            return f"<![CDATA[{text or ''}]]>"

        lines.append("  <item>")
        lines.append(f"    <title>{cdata(p.title)}</title>")
        lines.append(f"    <time_limit unit=\"s\">{cdata(str(tl))}</time_limit>")
        lines.append(f"    <memory_limit unit=\"mb\">{cdata(str(ml))}</memory_limit>")
        lines.append(f"    <description>{cdata(p.description or '')}</description>")
        lines.append(f"    <input>{cdata(p.input or '')}</input>")
        lines.append(f"    <output>{cdata(p.output or '')}</output>")
        lines.append(f"    <sample_input>{cdata(p.sample_input or '')}</sample_input>")
        lines.append(f"    <sample_output>{cdata(p.sample_output or '')}</sample_output>")
        for idx, t in enumerate(tests):
            name = f"test{idx + 1:03d}"
            lines.append(f"    <test_input name=\"{name}\">{cdata(t['input'])}</test_input>")
            lines.append(f"    <test_output name=\"{name}\">{cdata(t['output'])}</test_output>")
        lines.append(f"    <hint>{cdata(p.hint or '')}</hint>")
        lines.append(f"    <source>{cdata(p.source or '')}</source>")
        lines.append("    <front></front>")
        lines.append("    <rear></rear>")
        lines.append("    <bann></bann>")
        lines.append("    <credits></credits>")
        lines.append("  </item>")
    lines.append("</CSLfps>")
    return "\n".join(lines)


@router.get("/problems/{problem_id}/export")
def export_problem_xml(
    problem_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    """단일 문제를 CSLfps XML로 내보내기"""
    p = db.query(Problem).filter(Problem.problem_id == problem_id, Problem.defunct == "N").first()
    if not p:
        raise HTTPException(status_code=404, detail="문제를 찾을 수 없습니다")
    tests = _read_testcases(problem_id)
    xml_str = _build_xml([{"problem": p, "tests": tests}])
    filename = f"problem_{problem_id}.xml"
    return Response(
        content=xml_str.encode("utf-8"),
        media_type="application/xml",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/problems/import")
async def import_problems_xml(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(get_current_admin),
):
    """CSLfps XML 파일을 파싱하여 문제 일괄 등록"""
    content = await file.read()
    try:
        root = ET.fromstring(content.decode("utf-8"))
    except ET.ParseError as e:
        raise HTTPException(status_code=400, detail=f"XML 파싱 오류: {e}")

    # CSLfps 또는 oj-export 루트 모두 허용
    if root.tag not in ("CSLfps", "oj-export"):
        raise HTTPException(status_code=400, detail=f"지원하지 않는 XML 포맷입니다 (root: {root.tag})")

    items = root.findall("item")
    if not items:
        raise HTTPException(status_code=400, detail="문제 항목(item)이 없습니다")

    def get_text(el, tag: str) -> str:
        child = el.find(tag)
        return (child.text or "").strip() if child is not None else ""

    created = []
    for item in items:
        title = get_text(item, "title") or "제목 없음"
        try:
            tl_raw = get_text(item, "time_limit")
            time_limit = float(tl_raw) if tl_raw else 1.0
        except ValueError:
            time_limit = 1.0
        try:
            ml_raw = get_text(item, "memory_limit")
            memory_limit = int(float(ml_raw)) if ml_raw else 128
        except ValueError:
            memory_limit = 128

        p = Problem(
            title=title,
            description=get_text(item, "description") or None,
            input=get_text(item, "input") or None,
            output=get_text(item, "output") or None,
            sample_input=get_text(item, "sample_input") or None,
            sample_output=get_text(item, "sample_output") or None,
            hint=get_text(item, "hint") or None,
            source=get_text(item, "source") or None,
            time_limit=time_limit,
            memory_limit=memory_limit,
            in_date=now_kst(),
            defunct="N",
            accepted=0,
            submit=0,
            solved=0,
        )
        db.add(p)
        db.flush()  # problem_id 확보

        # 테스트 케이스 수집 (name 속성 기준 정렬)
        test_pairs: dict[str, dict] = {}
        for child in item:
            if child.tag == "test_input":
                name = child.get("name", f"_{len(test_pairs)}")
                test_pairs.setdefault(name, {})["input"] = (child.text or "").strip()
            elif child.tag == "test_output":
                name = child.get("name", f"_{len(test_pairs)}")
                test_pairs.setdefault(name, {})["output"] = (child.text or "").strip()

        tests = [
            {"input": v.get("input", ""), "output": v.get("output", "")}
            for v in test_pairs.values()
            if "input" in v
        ]
        if tests:
            _write_testcases(p.problem_id, tests)

        created.append({"problem_id": p.problem_id, "title": p.title, "test_count": len(tests)})

    db.commit()
    return {"imported": len(created), "problems": created}

